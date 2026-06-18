import logging
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from collectors.base import Job
from storage.db import Database
from storage.excel_writer import export_excel, import_manual_fields
from utils.hashing import create_job_hash


class StorageExcelTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.database = Database(self.root / "jobs.sqlite")
        self.excel = self.root / "job_radar.xlsx"
        self.settings = {
            "excel": {"minimum_export_score": 35, "backup_count": 2},
            "scheduler": {"run_times": ["09:00", "12:00", "16:00", "20:00"]},
        }
        self.logger = logging.getLogger("test")

    def tearDown(self):
        self.database.close()
        self.temporary.cleanup()

    def _job(self):
        job = Job(
            title="Rendering Engineer",
            company="Acme",
            location="Remote",
            url="https://example.com/jobs/1",
            source_name="Test",
            source_type="html",
            match_score=80,
            match_reason="Strong match",
        )
        job.job_hash = create_job_hash(job.title, job.company, job.url, job.location)
        return job

    def test_upsert_preserves_status_and_excel_round_trip(self):
        job = self._job()
        self.assertTrue(self.database.upsert_job(job, "2026-06-08T09:00:00-04:00"))
        self.database.add_run_log(
            {
                "run_time": "2026-06-08T09:00:00-04:00",
                "sources_checked": 1,
                "jobs_found": 1,
                "new_jobs_added": 1,
                "errors": [],
                "duration_seconds": 1.0,
            }
        )
        export_excel(self.database, self.excel, [], {}, self.settings, self.logger)

        workbook = load_workbook(self.excel)
        sheet = workbook["Jobs"]
        sheet["C2"] = "Applied"
        sheet["M2"] = "Applied on company site"
        workbook.save(self.excel)
        workbook.close()

        self.assertEqual(import_manual_fields(self.excel, self.database, self.logger), 1)
        job.description = "Updated description"
        self.assertFalse(self.database.upsert_job(job, "2026-06-08T12:00:00-04:00"))
        row = self.database.list_jobs(35)[0]
        self.assertEqual(row["status"], "Applied")
        self.assertEqual(row["notes"], "Applied on company site")

    def test_low_score_is_stored_but_not_exported(self):
        job = self._job()
        job.match_score = 10
        self.database.upsert_job(job)
        export_excel(self.database, self.excel, [], {}, self.settings, self.logger)
        workbook = load_workbook(self.excel, read_only=True)
        self.assertEqual(workbook["Jobs"].max_row, 1)
        workbook.close()

    def test_delete_jobs_containing_phrase_removes_matching_rows(self):
        job = self._job()
        job.description = (
            "Must be a U.S. Person due to required access to U.S. export controlled information or facilities"
        )
        self.database.upsert_job(job)
        removed = self.database.delete_jobs_containing(
            "Must be a U.S. Person due to required access to U.S. export controlled information or facilities"
        )
        self.assertEqual(removed, 1)
        self.assertEqual(self.database.list_jobs(0), [])

    def test_delete_jobs_by_hashes_removes_matching_rows(self):
        job = self._job()
        self.database.upsert_job(job)
        removed = self.database.delete_jobs_by_hashes([job.job_hash])
        self.assertEqual(removed, 1)
        self.assertEqual(self.database.list_jobs(0), [])
