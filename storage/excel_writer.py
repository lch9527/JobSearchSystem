from __future__ import annotations

import logging
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from storage.db import Database
from storage.models import ALLOWED_STATUSES, DEFAULT_STATUS
from utils.text import normalize_url


JOB_HEADERS = [
    "Date Found",
    "Last Seen",
    "Status",
    "Score",
    "Title",
    "Company",
    "Location",
    "Salary",
    "Employment Type",
    "Source",
    "URL",
    "Match Reason",
    "Notes",
    "Job Hash",
]


def import_manual_fields(path: Path, database: Database, logger: logging.Logger) -> int:
    if not path.exists():
        return 0
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "Jobs" not in workbook.sheetnames:
            workbook.close()
            return 0
        sheet = workbook["Jobs"]
        headers = {
            str(cell.value).strip(): index
            for index, cell in enumerate(next(sheet.iter_rows()), start=1)
            if cell.value
        }
        updated = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            status = _row_value(row, headers.get("Status")) or DEFAULT_STATUS
            notes = str(_row_value(row, headers.get("Notes")) or "")
            job_hash = _row_value(row, headers.get("Job Hash"))
            url = _row_value(row, headers.get("URL"))
            if database.update_manual_fields(
                str(job_hash) if job_hash else None,
                str(url) if url else None,
                str(status),
                notes,
            ):
                updated += 1
        workbook.close()
        return updated
    except Exception as exc:
        logger.warning("Could not import manual Excel fields: %s", exc)
        return 0


def export_excel(
    database: Database,
    path: Path,
    sources: list[dict[str, Any]],
    keyword_config: dict[str, Any],
    settings: dict[str, Any],
    logger: logging.Logger,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    excel_settings = settings.get("excel", {})
    minimum_score = int(excel_settings.get("minimum_export_score", 35))
    _backup_existing(path, int(excel_settings.get("backup_count", 20)), logger)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    _write_jobs(workbook, database.list_jobs(minimum_score))
    _write_sources(workbook, sources, database.source_checks())
    _write_keywords(workbook, keyword_config)
    _write_summary(workbook, database.status_counts(), settings)
    _write_run_logs(workbook, database.list_run_logs())

    temporary = path.with_name(f".{path.stem}.tmp{path.suffix}")
    try:
        workbook.save(temporary)
        temporary.replace(path)
        return path
    except PermissionError:
        fallback = path.with_name(
            f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}"
        )
        workbook.save(fallback)
        logger.warning("Excel file is busy; wrote fallback file %s", fallback)
        return fallback
    finally:
        if temporary.exists():
            temporary.unlink()


def _write_jobs(workbook: Workbook, jobs: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Jobs")
    sheet.append(JOB_HEADERS)
    for job in jobs:
        sheet.append(
            [
                job["date_found"],
                job["last_seen"],
                job["status"],
                job["match_score"],
                job["title"],
                job["company"],
                job["location"],
                job["salary"],
                job["employment_type"],
                job["source_name"],
                job["url"],
                job["match_reason"],
                job["notes"],
                job["job_hash"],
            ]
        )
        url_cell = sheet.cell(sheet.max_row, 11)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.style = "Hyperlink"

    status_list = ",".join(sorted(ALLOWED_STATUSES))
    validation = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=False)
    sheet.add_data_validation(validation)
    validation.add(f"C2:C1048576")
    sheet.column_dimensions["N"].hidden = True
    _format_sheet(sheet, widths=[22, 22, 14, 9, 38, 24, 24, 18, 18, 22, 55, 75, 40])


def _write_sources(
    workbook: Workbook,
    sources: list[dict[str, Any]],
    checks: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Sources")
    sheet.append(["Source Name", "Source Type", "URL", "Enabled", "Last Checked", "Notes"])
    for source in sources:
        name = source.get("name", "")
        url = source.get("url", "")
        if not url:
            token = source.get("company_slug") or source.get("board_token") or ""
            url = str(token)
        sheet.append(
            [
                name,
                source.get("type", ""),
                url,
                bool(source.get("enabled", False)),
                checks.get(name, {}).get("last_checked", ""),
                source.get("notes", ""),
            ]
        )
    _format_sheet(sheet, widths=[28, 15, 50, 12, 22, 60])


def _write_keywords(workbook: Workbook, config: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Keywords")
    sheet.append(["Keyword", "Type", "Weight", "Notes"])
    sections = [
        ("include_keywords", "include", "weight"),
        ("title_boost_keywords", "title_boost", "weight"),
        ("exclude_keywords", "exclude", "penalty"),
    ]
    for section, keyword_type, value_key in sections:
        for entry in config.get(section, []):
            sheet.append(
                [
                    entry.get("keyword", ""),
                    keyword_type,
                    entry.get(value_key, 0),
                    entry.get("notes", ""),
                ]
            )
    _format_sheet(sheet, widths=[30, 20, 12, 60])


def _write_summary(
    workbook: Workbook, counts: dict[str, int], settings: dict[str, Any]
) -> None:
    sheet = workbook.create_sheet("Summary")
    now = datetime.now().astimezone()
    rows = [
        ("Total Jobs", counts.get("Total", 0)),
        ("New Jobs", counts.get("New", 0)),
        ("Interested Jobs", counts.get("Interested", 0)),
        ("Applied Jobs", counts.get("Applied", 0)),
        ("Not Match Jobs", counts.get("Not Match", 0)),
        ("Rejected Jobs", counts.get("Rejected", 0)),
        ("Archived Jobs", counts.get("Archived", 0)),
        ("Last Run Time", now.isoformat(timespec="seconds")),
        ("Next Expected Run Time", _next_run(now, settings)),
    ]
    sheet.append(["Metric", "Value"])
    for row in rows:
        sheet.append(row)
    _format_sheet(sheet, widths=[28, 30])


def _write_run_logs(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Run_Log")
    sheet.append(
        [
            "Run Time",
            "Sources Checked",
            "Jobs Found",
            "New Jobs Added",
            "Errors",
            "Duration Seconds",
        ]
    )
    for row in rows:
        sheet.append(
            [
                row["run_time"],
                row["sources_checked"],
                row["jobs_found"],
                row["new_jobs_added"],
                row["errors"],
                row["duration_seconds"],
            ]
        )
    _format_sheet(sheet, widths=[24, 18, 15, 18, 80, 18])


def _format_sheet(sheet, widths: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _backup_existing(path: Path, keep: int, logger: logging.Logger) -> None:
    if not path.exists() or keep <= 0:
        return
    backup_dir = path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S_%f}{path.suffix}"
    try:
        shutil.copy2(path, backup)
        backups = sorted(backup_dir.glob(f"{path.stem}_*{path.suffix}"), reverse=True)
        for old_backup in backups[keep:]:
            old_backup.unlink()
    except OSError as exc:
        logger.warning("Could not back up Excel file: %s", exc)


def _next_run(now: datetime, settings: dict[str, Any]) -> str:
    run_times = settings.get("scheduler", {}).get("run_times", [])
    candidates = []
    for value in run_times:
        try:
            hour, minute = (int(part) for part in str(value).split(":", maxsplit=1))
            candidate = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if candidate <= now:
                candidate += timedelta(days=1)
            candidates.append(candidate)
        except (TypeError, ValueError):
            continue
    return min(candidates).isoformat(timespec="seconds") if candidates else ""


def _row_value(row: tuple, one_based_index: int | None):
    if not one_based_index or one_based_index > len(row):
        return None
    return row[one_based_index - 1]

